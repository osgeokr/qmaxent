"""Background worker: full QMaxent pipeline in correct logical order.

Order: Sample background → Annotate → Train → ROC → CV → Jackknife

The pipeline is built on the following references; each call site has
inline citations pointing to whichever item is most relevant.

    Backend:
        Anderson, C.B. (2023). elapid: Species distribution modeling
            tools for Python. JOSS, 8(84), 4930.

    Maxent algorithm and feature classes:
        Phillips, S.J., Anderson, R.P., & Schapire, R.E. (2006).
            Maximum entropy modeling of species geographic
            distributions. Ecological Modelling, 190, 231-259.
        Phillips, S.J., & Dudik, M. (2008). Modeling of species
            distributions with Maxent: new extensions and a
            comprehensive evaluation. Ecography, 31, 161-175.
        Phillips, S.J., Anderson, R.P., Dudik, M., Schapire, R.E., &
            Blair, M.E. (2017). Opening the black box: an open-source
            release of Maxent. Ecography, 40, 887-893.
        Elith, J., Phillips, S.J., Hastie, T., Dudik, M., Chee, Y.E.,
            & Yates, C.J. (2011). A statistical explanation of MaxEnt
            for ecologists. Diversity and Distributions, 17, 43-57.
        Merow, C., Smith, M.J., & Silander, J.A. (2013). A practical
            guide to MaxEnt for modeling species' distributions.
            Ecography, 36, 1058-1069.

    Optimization and statistical equivalence:
        Friedman, J., Hastie, T., & Tibshirani, R. (2010).
            Regularization paths for generalized linear models via
            coordinate descent. JSS, 33(1), 1-22.
        Fithian, W., & Hastie, T. (2013). Finite-sample equivalence in
            statistical models for presence-only data. Annals of
            Applied Statistics, 7(4), 1917-1939.
        Radosavljevic, A., & Anderson, R.P. (2014). Making better
            Maxent models of species distributions: complexity,
            overfitting and evaluation. Journal of Biogeography, 41,
            629-643.

    Sample selection bias:
        Phillips, S.J., Dudik, M., Elith, J., Graham, C.H., Lehmann,
            A., Leathwick, J., & Ferrier, S. (2009). Sample selection
            bias and presence-only distribution models. Ecological
            Applications, 19, 181-197.

    Evaluation and spatial cross-validation:
        Pearson, R.G., Raxworthy, C.J., Nakamura, M., & Peterson, A.T.
            (2007). Predicting species distributions from small
            numbers of occurrence records. Journal of Biogeography,
            34, 102-117.
        Lobo, J.M., Jimenez-Valverde, A., & Real, R. (2008). AUC: a
            misleading measure of the performance of predictive
            distribution models. Global Ecology and Biogeography, 17,
            145-151.
        Muscarella, R., et al. (2014). ENMeval: an R package for
            conducting spatially independent evaluations and
            estimating optimal model complexity. MEE, 5, 1198-1205.
        Roberts, D.R., et al. (2017). Cross-validation strategies for
            data with temporal, spatial, hierarchical, or phylogenetic
            structure. Ecography, 40, 913-929.
        Valavi, R., Elith, J., Lahoz-Monfort, J.J., & Guillera-Arroita,
            G. (2019). blockCV: an R package for generating spatially
            or environmentally separated folds. MEE, 10, 225-232.
        Ploton, P., et al. (2020). Spatial validation reveals poor
            predictive performance of large-scale ecological mapping
            models. Nature Communications, 11, 4540.
"""

import os
import traceback

import numpy as np
from qgis.PyQt.QtCore import QThread, pyqtSignal

from ..core.maxent_rules import auto_feature_types
from ..i18n import tr


class MaxentWorker(QThread):
    progress = pyqtSignal(int, str)
    log = pyqtSignal(str)
    finished = pyqtSignal(bool, str, dict)

    def __init__(self, config: dict, parent=None):
        super().__init__(parent)
        self._cfg = config
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    # -----------------------------------------------------------------------
    # Entry point
    # -----------------------------------------------------------------------

    def run(self):
        try:
            # Disable tqdm's daemon monitor thread before any code path
            # imports tqdm via elapid (or anywhere else). See the
            # extensive comment in projection_worker.py — same crash
            # signature, same fix. We do this in every QMaxent worker
            # that touches elapid because a single instantiation of
            # tqdm.tqdm anywhere starts the monitor for the entire
            # process, and we want that to never happen.
            try:
                import tqdm as _tqdm

                _tqdm.tqdm.monitor_interval = 0
                try:
                    mon = getattr(_tqdm.tqdm, "monitor", None)
                    if mon is not None and hasattr(mon, "was_killed"):
                        mon.was_killed.set()
                        _tqdm.tqdm.monitor = None
                except Exception:
                    pass
            except Exception:
                pass

            result = self._execute()
            self.finished.emit(True, "Done", result)
        except InterruptedError:
            self.finished.emit(False, "Cancelled", {})
        except Exception as e:
            self.finished.emit(False, f"{e}\n{traceback.format_exc()}", {})

    def _check(self, pct: int, msg: str):
        # Update the progress bar only; the 6-block log structure
        # already announces each phase via explicit header lines.
        # Emitting progress messages into the log too would produce
        # noisy duplicates ("Training MaxentModel..." right next to
        # the "─── 2. Full-data model fit ───" header).
        if self._cancelled:
            raise InterruptedError("cancelled")
        self.progress.emit(pct, msg)

    # -----------------------------------------------------------------------
    # Pipeline
    # -----------------------------------------------------------------------

    def _execute(self) -> dict:
        import pandas as pd
        from sklearn.metrics import roc_auc_score
        from sklearn.metrics import roc_curve as _roc_curve

        from ..bridge import elapid_bridge as eb

        cfg = self._cfg
        presence_gdf = cfg["presence_gdf"]
        raster_paths = cfg["raster_paths"]
        feature_names = cfg["feature_names"]
        # 0-based indices of variables flagged categorical by the user in
        # the ① Data tab. Following the dismo / SDMtune / ENMeval / maxent.jar
        # convention, the user — not the plugin — declares which rasters are
        # categorical (e.g. land cover, biome, soil type). elapid then
        # one-hot encodes these via CategoricalTransformer and applies the
        # separate beta_categorical regularisation scaler (Anderson 2023;
        # Phillips et al. 2017).
        categorical_indices = list(cfg.get("categorical_indices", []) or [])
        # ─── 1. Data preparation ──────────────────────────────────────────
        self.log.emit("─── 1. Data preparation ───")
        if categorical_indices:
            cat_names = ", ".join(feature_names[i] for i in categorical_indices)
            self.log.emit(f"  Categorical variables: {cat_names}")
        # Default of 10,000 background points follows Phillips & Dudik (2008),
        # which has since become the de-facto standard for Maxent SDMs
        # (also used by maxent.jar and ENMeval). Smaller values may be
        # appropriate for very small study areas; larger values rarely add
        # information once the environmental space is well sampled.
        n_background = cfg.get("n_background", 10000)

        # Auto-create parent directories for any configured save paths
        # so the user can type a fresh output path (e.g. a new
        # qmaxent_output/ folder) without first running mkdir. This
        # prevents the cryptic FileNotFoundError that the elapid
        # save_object call would otherwise raise mid-run, after the
        # full training has already completed.
        for _key in ("output_model", "output_xlsx"):
            _p = cfg.get(_key)
            if _p:
                _d = os.path.dirname(_p)
                if _d:
                    try:
                        os.makedirs(_d, exist_ok=True)
                    except OSError as _e:
                        # Surface the failure now instead of mid-save
                        # (e.g. permission denied, path-too-long on
                        # Windows). The user can fix the path before
                        # waiting through training.
                        raise RuntimeError(
                            f"Could not create output directory '{_d}': {_e}"
                        ) from _e

        # Compatibility check (informational; not fatal).
        ok, ver, msg = eb.check_elapid_version()
        self.log.emit(f"  {msg}")

        # ── 0a. Raster consistency safety check (fail-fast) ──────────────
        # Following the explicit-pre-harmonization workflow used by
        # biomod2 (BIOMOD_FormatingData) and SDMSelect (Prepare_r_multi),
        # QMaxent does not silently reproject mismatched rasters at
        # train time. The user is expected to run "Check Raster
        # Consistency" → "Harmonize to Folder…" from the ① Data tab
        # before pressing Run. We still verify here as a safety net:
        # if the user pressed Run on a mismatched set anyway, we abort
        # with a clear, actionable error rather than producing a model
        # whose validation does not match its training configuration.
        from ..bridge.raster_bridge import (
            check_raster_consistency,
            get_raster_crs,
        )

        consistency = check_raster_consistency(raster_paths)
        if not consistency.get("is_consistent", False):
            mismatches = []
            if not consistency["crs_uniform"]:
                mismatches.append("CRS")
            if not consistency["extent_uniform"]:
                mismatches.append("extent")
            if not consistency["resolution_uniform"]:
                mismatches.append("resolution")
            raise RuntimeError(
                tr(
                    "Environmental rasters do not share a common grid "
                    '({mismatches} differ). Run "Check Raster Consistency" '
                    'and "Harmonize to Folder…" in the ① Data tab to '
                    "align them before training."
                ).format(mismatches=", ".join(mismatches))
            )
        else:
            self.log.emit("  Rasters share grid (CRS, extent, resolution). ✓")

        # ── 0. CRS reconciliation ────────────────────────────────────────
        # Spatial CV strategies (Geographic K-Fold, Buffered LOO) operate
        # in the units of the presence-layer CRS, and elapid samples
        # background points in the raster CRS. A silent mismatch produces
        # nonsensical distances (Roberts et al. 2017). Reproject the
        # presence GeoDataFrame to the first raster's CRS when needed.
        raster_crs = get_raster_crs(raster_paths[0])
        if raster_crs is not None and presence_gdf.crs is not None:
            try:
                same = presence_gdf.crs.to_string() == str(raster_crs)
            except Exception:
                # rasterio / pyproj raise a variety of errors when the
                # raster's CRS is malformed or not WKT-serialisable;
                # treat any failure as "different CRS" so we trigger
                # the reproject path on the safe side.
                same = False
            if not same:
                self.log.emit(
                    f"  ⚠ CRS mismatch — presence={presence_gdf.crs}, "
                    f"raster={raster_crs}; reprojecting presences."
                )
                presence_gdf = presence_gdf.to_crs(raster_crs)
        elif raster_crs is None:
            self.log.emit("  ⚠ Could not read raster CRS; skipping CRS check.")
        elif presence_gdf.crs is None:
            self.log.emit("  ⚠ Presence layer has no CRS; assuming it matches raster.")

        # ── 1. Background sampling ────────────────────────────────────────
        # When a bias raster is supplied we sample background points with
        # probability proportional to bias raster values, correcting for
        # spatial sampling effort (Phillips et al. 2009). Otherwise points
        # are drawn uniformly from the first raster's valid extent.
        self._check(5, f"Sampling {n_background:,} background points...")
        if cfg.get("bias_path"):
            bg_gs = eb.sample_bias_file(cfg["bias_path"], count=n_background)
        else:
            bg_gs = eb.sample_raster(raster_paths[0], count=n_background)
        # (No log emit here — the consolidated "Presence: N, Background: N"
        # line below covers both counts in one place.)

        # ── 2. Annotation ────────────────────────────────────────────────
        self._check(12, "Extracting raster covariates for presence points...")
        presence_ann = eb.annotate(presence_gdf, raster_paths, labels=feature_names, quiet=True)
        self._check(20, "Extracting raster covariates for background points...")
        background_ann = eb.annotate(bg_gs, raster_paths, labels=feature_names, quiet=True)

        presence_clean = presence_ann.dropna(subset=feature_names).reset_index(drop=True)
        background_clean = background_ann.dropna(subset=feature_names).reset_index(drop=True)
        n_presence = len(presence_clean)
        n_background = len(background_clean)
        self.log.emit(f"  → Presence: {n_presence}, Background: {n_background:,}")

        # Minimum-sample-size guards. After dropna it is possible to be
        # left with too few records to fit any meaningful model — for
        # example when most presences fall on raster NoData cells. We
        # check this explicitly and fail with a clear message rather
        # than letting elapid/sklearn raise a cryptic numerical error
        # several steps downstream. Thresholds follow the SDM literature
        # (van Proosdij et al. 2016 / Wisz et al. 2008): below n=5
        # presences a presence-only model is not informative; below
        # n=100 background points the maxent likelihood is unstable.
        if n_presence == 0:
            raise ValueError(
                tr(
                    "No valid presence points after removing rows with NoData "
                    "in the environmental rasters. Check that your presence "
                    "layer overlaps the raster extents and that the rasters "
                    "have valid values at the presence locations."
                )
            )
        if n_presence < 5:
            raise ValueError(
                tr(
                    "Too few presence points after removing NoData "
                    "(n={n} < 5). Maxent is not reliable below 5 records; "
                    "please collect more presences or expand the study area."
                ).format(n=n_presence)
            )
        if n_background < 100:
            raise ValueError(
                tr(
                    "Too few background points after removing NoData "
                    "(n={n} < 100). Increase the background sample count or "
                    "check the raster coverage of the study area."
                ).format(n=n_background)
            )

        if cfg.get("add_to_bg", True):
            # `addsamplestobackground=True` is the canonical Maxent default
            # (Phillips et al. 2006; 2017): presence points are also
            # included in the background sample so that the fitted density
            # is consistent over the full study area.
            background_clean = pd.concat([background_clean, presence_clean], ignore_index=True)

        # Feature types
        if cfg.get("feature_types_auto", True):
            feature_types = auto_feature_types(n_presence)
        else:
            # Fallback (only used when cfg lacks "feature_types"): full
            # LQPHT — same default the UI seeds in Manual selection
            # (Phillips et al. 2017's standard Maxent feature set).
            feature_types = cfg.get(
                "feature_types", ["linear", "quadratic", "hinge", "product", "threshold"]
            )
        ft_origin = "auto-rule" if cfg.get("feature_types_auto", True) else "manual"
        self.log.emit(f"  → Feature types: {feature_types}  (n_presence={n_presence}, {ft_origin})")

        # Merge (presence FIRST — index order relied on by CV logic)
        presence_clean = presence_clean.assign(**{"class": 1})
        background_clean = background_clean.assign(**{"class": 0})
        merged = pd.concat(
            [
                presence_clean[feature_names + ["class"]],
                background_clean[feature_names + ["class"]],
            ],
            ignore_index=True,
        )
        x = merged[feature_names].values.astype(np.float32)
        y = merged["class"].values.astype(np.uint8)

        # ── 2b. Optional sample weights ─────────────────────────────────────
        # When the user opts in to spatial bias correction we use elapid's
        # distance_weights to assign lower weights to spatially clustered
        # presence points and higher weights to isolated ones. Following
        # Phillips et al. (2009), the weights are computed on presences
        # only; background points retain weight 1.0. This implements
        # exactly what dismo, ENMeval, biomod2 leave to the user as an
        # opt-in correction step.
        sample_weight = None
        if cfg.get("distance_weights", False):
            try:
                import elapid as _elapid

                pres_geom = presence_clean.geometry
                w_pres = _elapid.distance_weights(pres_geom)
                # background points keep weight 1.0
                w_bg = np.ones(len(background_clean), dtype=np.float32)
                sample_weight = np.concatenate([np.asarray(w_pres, dtype=np.float32), w_bg])
                self.log.emit(
                    f"  → Sample bias correction: distance-based weights "
                    f"applied to {len(pres_geom)} presences "
                    f"(Phillips 2009; elapid distance_weights)"
                )
            except Exception as e:
                # Fallback: warn and proceed without weights rather than
                # silently silently dropping the user's explicit opt-in.
                self.log.emit(
                    f"  ⚠ distance_weights failed ({e}); proceeding without sample weights"
                )
                sample_weight = None

        # ── 3. Train ──────────────────────────────────────────────────────
        # Defaults are documented in bridge.elapid_bridge.make_maxent_model.
        # Briefly:
        #   beta_multiplier       Phillips & Dudik 2008; Radosavljevic 2014
        #   transform="cloglog"   Phillips et al. 2017 (recommended default)
        #   n_hinge_features=50   maxnet's documented default (Phillips
        #   n_threshold_features=50  Java app behaviour, replicated by the
        #                            maxnet R package: hinge() and
        #                            thresholds() both take nknots=50).
        #                            Each variable contributes 2*nknots-2
        #                            hinge/threshold features. elapid's
        #                            own default is 10; we override to 50
        #                            so QMaxent matches maxnet.
        #   use_lambdas/n_lambdas Friedman, Hastie & Tibshirani 2010 (glmnet)
        #   class_weights=100     Fithian & Hastie 2013 (IWLR equivalence)
        # ─── 2. Full-data model fit ───────────────────────────────────────
        self.log.emit("")
        self.log.emit("─── 2. Full-data model fit ───")
        self._check(32, "Training MaxentModel...")
        self.log.emit("  Training MaxentModel on full data...")
        model = eb.make_maxent_model(
            feature_types=feature_types,
            beta_multiplier=cfg.get("beta_mult", 1.0),
            transform=cfg.get("transform", "cloglog"),
            n_hinge_features=cfg.get("n_hinge", 50),
            n_threshold_features=cfg.get("n_threshold", 50),
            use_lambdas="last",
            n_lambdas=200,
            class_weights=100,
            use_sklearn=True,
        )
        model.fit(
            x,
            y,
            sample_weight=sample_weight,
            categorical=categorical_indices or None,
            labels=feature_names,
        )
        # (No "Model training complete" log line here — the Training AUC
        # emit below is itself the completion signal.)

        # Attach metadata (response curves / projection)
        x_presence = x[y == 1]

        # Build per-feature hold-out values used in response-curve plots.
        # Continuous variables: arithmetic mean of the presence sample —
        #   this is the standard maxent.jar / dismo convention (sweep
        #   one variable, hold the others at their mean).
        # Categorical variables: most frequent value (mode) of the
        #   presence sample. The mean of integer category codes (e.g.
        #   biome ∈ {1,2,4,7,8,13}) is a non-integer like 1.8275, which
        #   sklearn's OneHotEncoder rejects with "unknown categories"
        #   because it never saw 1.8275 during training. Mode keeps the
        #   hold-out inside the encoder's vocabulary and matches how
        #   biomod2 / SDMtune build their response-curve background.
        cat_set = set(categorical_indices or [])
        samplemeans = {}
        for i, nm in enumerate(feature_names):
            col = x_presence[:, i]
            if i in cat_set and len(col):
                # Round to the nearest integer first so float-typed
                # categorical rasters (uint8 read as float32 by elapid)
                # collapse back onto their original codes, then take
                # the mode. np.bincount needs non-negative ints, so
                # we shift by min when needed.
                ints = np.rint(col).astype(np.int64)
                lo = int(ints.min())
                counts = np.bincount(ints - lo)
                mode_val = int(counts.argmax() + lo)
                samplemeans[nm] = float(mode_val)
            else:
                samplemeans[nm] = float(col.mean()) if len(col) else 0.0

        # Compute the canonical SDM threshold values up-front so the
        # Priority Sites for Survey tab can show them instantly when
        # the user selects a method (Pearson 2007 MTP / T10; Liu 2013
        # MaxSSS). These are derived from the model's predictions at
        # the training presence and background samples — exactly the
        # interpretation in the original references.
        try:
            train_pres_pred = model.predict(x[y == 1])
            train_bg_pred = model.predict(x[y == 0])
            train_pres_pred = np.asarray(train_pres_pred, dtype=float).ravel()
            train_bg_pred = np.asarray(train_bg_pred, dtype=float).ravel()
            thresholds = {
                "MTP": float(train_pres_pred.min()),
                "T10": float(np.percentile(train_pres_pred, 10)),
            }
            # MaxSSS — exhaustive scan over candidate cut-offs, picking
            # the one that maximises sensitivity + specificity.
            cands = np.unique(np.concatenate([train_pres_pred, train_bg_pred]))
            best_t, best_score = float(train_pres_pred.min()), -np.inf
            for t in cands:
                sens = float((train_pres_pred >= t).mean())
                spec = float((train_bg_pred < t).mean())
                if sens + spec > best_score:
                    best_score = sens + spec
                    best_t = float(t)
            thresholds["MaxSSS"] = best_t
        except Exception:
            # Predictions can fail (e.g., model wasn't fully fitted on
            # this dataset). The Priority Sites tab will fall back to
            # computing the threshold from the prediction raster
            # itself — slower but functionally identical.
            thresholds = {}

        # Capture the actual category codes seen in training for each
        # categorical variable. The response-curve plot needs this so
        # it can sweep ONLY known codes — sklearn's OneHotEncoder
        # rejects unknowns during predict() with the cryptic message
        # "Found unknown categories ... in column 0 during transform".
        # We collect from the combined presence + background design
        # matrix (i.e. x), which is exactly what fit() saw — not just
        # from presences, since some codes may appear only in the
        # background sample but still be part of the trained encoder.
        categorical_levels = {}
        for j in categorical_indices or []:
            nm = feature_names[j]
            col = x[:, j]
            if len(col):
                codes = sorted({int(round(float(v))) for v in col})
                categorical_levels[nm] = codes

        model._qmaxent_meta = {
            "feature_names": feature_names,
            "feature_types": feature_types,
            "categorical_indices": list(categorical_indices),
            "samplemeans": samplemeans,
            "transform": cfg.get("transform", "cloglog"),
            "n_presence": int(y.sum()),
            "n_background": int((y == 0).sum()),
            "beta_multiplier": cfg.get("beta_mult", 1.0),
            "distance_weights": bool(cfg.get("distance_weights", False)),
            "varmin": {nm: float(x[:, i].min()) for i, nm in enumerate(feature_names)},
            "varmax": {nm: float(x[:, i].max()) for i, nm in enumerate(feature_names)},
            "categorical_levels": categorical_levels,
            "thresholds": thresholds,
            # Always record the *original* user-supplied raster paths,
            # not the temporary harmonized copies. The harmonized
            # directory is wiped when the worker finishes; a saved .pkl
            # that referenced /tmp/qmaxent_harmonized_*/... would be
            # broken on the next QGIS session.
            # raster_paths recorded for downstream projection. After
            # removing the auto-harmonize codepath these are simply the
            # paths the user supplied — there is no longer a distinction
            # between "user-supplied" and "internally harmonized".
            "raster_paths": raster_paths,
        }

        # ── 4. ROC curve (training data) ──────────────────────────────────
        # Must be computed immediately after training, before CV retraining.
        # Note: the AUC reported here is *training* AUC. It is shown only
        # as a within-sample reference for the ROC curve and is known to
        # overestimate generalization performance for presence-only SDMs
        # (Lobo, Jimenez-Valverde & Real 2008). Held-out generalization is
        # reported by the spatial CV step (5) and the jackknife step (6).
        self._check(40, "Computing ROC curve...")
        roc_fpr, roc_tpr, train_auc = [], [], None
        try:
            preds = model.predict(x)
            fpr, tpr, _ = _roc_curve(y, preds)
            train_auc = float(roc_auc_score(y, preds))
            roc_fpr = fpr.tolist()
            roc_tpr = tpr.tolist()
            self.log.emit(f"  → Training AUC = {train_auc:.4f}")
        except Exception as e:
            self.log.emit(f"  → ROC computation skipped: {e}")

        # Save model after the headline metric so the log reads
        # "Training AUC = ...; Model saved: ..." in natural order.
        if cfg.get("output_model"):
            eb.save_object(model, cfg["output_model"])
            self.log.emit(f"  → Model saved: {cfg['output_model']}")

        # Start building result dict — ROC data goes in right away
        result = {
            "model": model,
            "meta": model._qmaxent_meta,
            "x": x,
            "y": y,
            "presence_clean": presence_clean,
            "feature_names": feature_names,
            "roc_fpr": roc_fpr,
            "roc_tpr": roc_tpr,
            "full_auc": train_auc,
        }

        # ── 5. Cross-validation (generalization check) ────────────────────
        cv_method = cfg.get("cv_method", 0)
        first_fold = None  # (train_mask, test_mask) for jackknife reuse
        if cv_method > 0:
            # ─── 3. Cross-validation ──────────────────────────────────
            # Header text adapts per-method: K-Fold methods report
            # "n=N folds", Checkerboard reports "single split", BLOO
            # reports the buffer distance (its splits are determined
            # by buffer + n_presence, not a fold count).
            seed_cfg = cfg.get("random_seed")
            seed_str = f", seed={int(seed_cfg)}" if seed_cfg is not None else ", seed=random"
            if cv_method in (1, 2):
                cv_name_map = {1: "Geographic K-Fold", 2: "Random K-Fold"}
                cv_name = cv_name_map[cv_method]
                n_folds_cfg = int(cfg.get("n_folds", 5))
                header_detail = f"n={n_folds_cfg}{seed_str}"
            elif cv_method == 3:
                cv_name = "Checkerboard"
                grid_m = cfg.get("grid_size", 50000.0)
                header_detail = f"single split, grid={int(grid_m)} m"
            elif cv_method == 4:
                cv_name = "Buffered LOO"
                buf_m = cfg.get("buffer_dist", 50000)
                header_detail = f"buffer={int(buf_m)} m{seed_str}"
            else:
                cv_name = f"method {cv_method}"
                header_detail = ""
            self.log.emit("")
            self.log.emit(f"─── 3. Cross-validation ({cv_name}, {header_detail}) ───")
            self._check(50, "Running cross-validation...")
            (cv_aucs, first_fold, cv_roc_fpr, cv_roc_tpr) = self._cross_validate(
                x,
                y,
                presence_clean,
                feature_names,
                feature_types,
                cfg,
                cv_method,
                categorical_indices=categorical_indices,
                sample_weight=sample_weight,
            )
            result["cv_aucs"] = cv_aucs
            result["cv_roc_fpr_list"] = cv_roc_fpr
            result["cv_roc_tpr_list"] = cv_roc_tpr
            if cv_aucs:
                valid = [a for a in cv_aucs if not np.isnan(a)]
                if valid:
                    # Single-split methods (Checkerboard, BLOO) produce
                    # one AUC, so "mean ± std" would be misleading
                    # (std is trivially 0 for one value). Report just
                    # the AUC in that case.
                    if cv_method in (3, 4) or len(valid) == 1:
                        self.log.emit(f"  → CV AUC (held-out) = {np.mean(valid):.4f}")
                    else:
                        self.log.emit(
                            f"  → CV AUC (held-out, mean ± std) = "
                            f"{np.mean(valid):.4f} ± {np.std(valid):.4f}"
                        )
                else:
                    # All folds returned NaN. Show the user a clear message
                    # rather than reporting "nan ± nan", which is what
                    # np.mean([]) / np.std([]) would emit.
                    self.log.emit(
                        "  ⚠ CV failed: no valid fold (all folds returned NaN). "
                        "This usually means too few presences in each fold or "
                        "that the model could not be trained on a fold."
                    )

        # ── 6. Jackknife (variable importance, uses full model) ───────────
        if cfg.get("do_jackknife", True):
            # ─── 4. Jackknife variable importance ──────────────────────
            # The "reuse CV" path: pass cv_aucs into the jackknife so
            # its full-model reference matches the headline CV AUC
            # exactly (no second retrain, no second fold-split). Both
            # values are anchored to the same partitioning the user
            # configured in ② Parameters.
            # ─── Issue A (v0.1.7 final): BLOO opts OUT of cv_aucs reuse.
            # BLOO produces a single pooled AUC across N pseudo-folds —
            # methodologically distinct from K-Fold's per-fold AUCs.
            # Using it as the jackknife "full-model reference" while
            # per-variable AUCs are evaluated on a separate 25%
            # hold-out (because first_fold is None for BLOO) would
            # mix two non-comparable measurement schemes in the same
            # chart. Cleaner: BLOO's pooled AUC stays in § 5; the
            # jackknife reverts to its 25%-hold-out path (same as
            # cv_method=0 None), keeping reference and per-variable
            # AUCs anchored to the same fold.
            # Checkerboard (cv_method=3) keeps cv_aucs reuse because
            # its single train/test split IS the first_fold — so
            # jackknife reuses the exact same partition.
            if cv_method in (1, 2, 3):
                cv_aucs_for_jk = result.get("cv_aucs", None)
            else:
                cv_aucs_for_jk = None
            # ─── Issue B (v0.1.7 final): per-method header wording.
            # "across the same N folds" is only natural for actual
            # K-Fold methods. Checkerboard is a single spatial
            # split (Muscarella 2014), and BLOO / None fall back to
            # a random 25% hold-out.
            self.log.emit("")
            if cv_method in (1, 2) and cv_aucs_for_jk:
                n_folds_for_jk = len([a for a in cv_aucs_for_jk if not np.isnan(a)])
                self.log.emit(
                    f"─── 4. Jackknife variable importance "
                    f"(across the same {n_folds_for_jk} folds) ───"
                )
            elif cv_method == 3 and cv_aucs_for_jk:
                self.log.emit(
                    "─── 4. Jackknife variable importance (on the same checkerboard split) ───"
                )
            else:
                # cv_method == 0 (None) or 4 (BLOO).
                self.log.emit("─── 4. Jackknife variable importance (25% hold-out) ───")
            self._check(75, "Running jackknife variable importance...")
            # Inject presence_clean into cfg so the per-fold loop can
            # use Geographic K-Fold (which needs coordinates) without
            # changing the public signature of _jackknife. We use a
            # leading-underscore key to keep it clearly internal.
            cfg["_presence_clean"] = presence_clean
            jk_results, jk_full_train, jk_full_test = self._jackknife(
                model,
                x,
                y,
                feature_names,
                feature_types,
                cfg,
                first_fold,
                categorical_indices=categorical_indices,
                sample_weight=sample_weight,
                cv_aucs=cv_aucs_for_jk,
            )
            cfg.pop("_presence_clean", None)
            result["jackknife_results"] = jk_results
            # Store the jackknife's full-model reference AUCs so the
            # importance chart can plot "Full model" with the exact
            # value, not a round-trip-error reconstruction.
            result["jk_full_train_auc"] = jk_full_train
            result["jk_full_test_auc"] = jk_full_test
        else:
            # Without jackknife the loop above never runs, so we need an
            # intermediate checkpoint to keep the progress bar moving
            # smoothly (75% → 95% → 100%) instead of jumping directly
            # from 75% to 100% and looking frozen.
            self._check(85, "Finalizing model...")

        # ── 7. Permutation importance (added v0.1.3) ─────────────────────
        # Reports the AUC drop when each feature's values are randomly
        # shuffled — the same metric reported by maxent.jar as
        # "permutation importance" (Phillips et al. 2006, 2017). Computed
        # on the held-out test fold when a CV first_fold exists,
        # otherwise on the full training matrix. Result is normalized
        # to percentage of total to match maxent.jar's output
        # convention and enable direct ranking comparison with
        # published Maxent SDM studies.
        #
        # Note (Strobl et al. 2007; Hooker & Mentch 2019): permutation
        # importance can underestimate the importance of features
        # correlated with other features in the model. We report the
        # raw values without correction and document this caveat in
        # the UI panel; users should interpret jackknife and
        # permutation importance together rather than relying on
        # either alone.
        if cfg.get("do_permutation", True):
            self._check(90, "Computing permutation importance...")
            try:
                # Use the held-out test fold for an honest estimate
                # when CV is configured. Falls back to the full design
                # matrix when no CV partition is available.
                if first_fold is not None:
                    _, te_mask = first_fold
                    x_pi, y_pi = x[te_mask], y[te_mask]
                    pi_source = "held-out test set"
                else:
                    x_pi, y_pi = x, y
                    pi_source = "training set"

                n_repeats = int(cfg.get("permutation_repeats", 10))
                # ─── 5. Permutation importance ─────────────────────────
                self.log.emit("")
                self.log.emit(
                    f"─── 5. Permutation importance "
                    f"(sklearn, n_repeats={n_repeats}, "
                    f"evaluated on {pi_source}) ───"
                )
                pi_seed = cfg.get("random_seed")
                importances = eb.permutation_importance_scores(
                    model,
                    x_pi,
                    y_pi,
                    n_repeats=n_repeats,
                    n_jobs=1,
                    random_state=pi_seed,
                )
                # importances shape: (n_features, n_repeats)
                pi_mean = importances.mean(axis=1)
                pi_std = importances.std(axis=1)

                # Normalize as percentage of total (matches maxent.jar's
                # permutation importance convention). Guard against
                # all-zero importances (e.g. perfectly random model).
                total = float(pi_mean.sum())
                if total > 0:
                    pi_pct = 100.0 * pi_mean / total
                else:
                    pi_pct = np.zeros_like(pi_mean)

                permutation_results = []
                for i, nm in enumerate(feature_names):
                    permutation_results.append(
                        {
                            "variable": nm,
                            "importance_mean": float(pi_mean[i]),
                            "importance_std": float(pi_std[i]),
                            "importance_pct": float(pi_pct[i]),
                        }
                    )
                # Sort by importance_mean descending so the chart and
                # XLSX both show the most important variable on top.
                permutation_results.sort(key=lambda r: -r["importance_mean"])
                result["permutation_results"] = permutation_results
                result["permutation_source"] = pi_source
                result["permutation_n_repeats"] = n_repeats
                self.log.emit(
                    f"  → Computed for {len(feature_names)} variables "
                    "(see Results > Permutation Importance tab)"
                )
            except Exception as e:
                # Permutation can fail when the held-out set is
                # degenerate (e.g. all-presence). Log and continue
                # rather than aborting the whole run — jackknife
                # already provides a variable-importance signal.
                self.log.emit(f"  → Permutation importance skipped: {e}")
                result["permutation_results"] = []

        # Persist academic results into the model meta so they survive
        # a save → load cycle. Without this, reopening a .pkl would
        # show empty Jackknife / ROC charts; the user would have to
        # retrain to see them again. We namespace them under
        # ``academic_results`` to keep the top-level meta dict clean.
        try:
            model._qmaxent_meta["academic_results"] = {
                "roc_fpr": list(result.get("roc_fpr", [])),
                "roc_tpr": list(result.get("roc_tpr", [])),
                "full_auc": result.get("full_auc"),
                # The jackknife's full-model AUCs (computed on the
                # first CV fold's train/test split). Used by the
                # importance chart so "Full model" shows the exact
                # AUC instead of an error-accumulating reconstruction
                # of without + drop.
                "jk_full_train_auc": result.get("jk_full_train_auc"),
                "jk_full_test_auc": result.get("jk_full_test_auc"),
                "cv_aucs": list(result.get("cv_aucs", [])),
                # Per-fold ROC traces so the load-model UX can replot
                # the academic-style ROC panel (each fold + mean +
                # std band) instead of just training ROC + a CV-AUC
                # number. Each entry is a list of floats.
                "cv_roc_fpr_list": [list(f) for f in result.get("cv_roc_fpr_list", [])],
                "cv_roc_tpr_list": [list(t) for t in result.get("cv_roc_tpr_list", [])],
                # Jackknife rows are dicts of plain Python types (str
                # for variable name, float for AUC). Copy defensively
                # so we don't pickle live numpy scalars.
                "jackknife_results": [
                    {k: (float(v) if isinstance(v, (int, float)) else v) for k, v in row.items()}
                    for row in result.get("jackknife_results", [])
                ],
                # Permutation importance rows (added v0.1.3). Same
                # defensive-copy pattern as jackknife — keeps the
                # .pkl portable across machines and free of numpy
                # scalar references.
                "permutation_results": [
                    {k: (float(v) if isinstance(v, (int, float)) else v) for k, v in row.items()}
                    for row in result.get("permutation_results", [])
                ],
                "permutation_source": result.get("permutation_source"),
                "permutation_n_repeats": result.get("permutation_n_repeats"),
            }
            # Re-save the .pkl now that academic_results is populated.
            # The first save_object call (right after model.fit, before
            # ROC / CV / jackknife had run) wrote a partial model to
            # disk; reloading it later would produce empty Results-tab
            # charts. Saving again here folds the analytics into the
            # file. We keep the earlier save call too so an interrupted
            # run (user clicks Stop after fit but before jackknife
            # completes) still leaves a usable — though chart-less —
            # .pkl on disk rather than nothing at all.
            if cfg.get("output_model"):
                try:
                    eb.save_object(model, cfg["output_model"])
                except Exception as e:
                    self.log.emit(
                        f"  → Could not re-save model with academic "
                        f"results ({type(e).__name__}); charts will "
                        f"not survive a reload of this .pkl."
                    )
        except Exception:
            # Failure here would silently cripple the load-model UX
            # (no charts after reload), but should never block training.
            # We log it so a developer running with --verbose can see it.
            self.log.emit("  → Could not persist academic results to meta.")

        # Save XLSX (multi-sheet styled report — see _save_xlsx).
        if cfg.get("output_xlsx") and (
            "jackknife_results" in result or "cv_aucs" in result or "permutation_results" in result
        ):
            # ─── 6. Save results ──────────────────────────────────────
            self.log.emit("")
            self.log.emit("─── 6. Save results ───")
            self._save_xlsx(result, cfg)
            result["xlsx_path"] = cfg["output_xlsx"]

        # No "Done" log line — the progress bar reaching 100% and the
        # main thread's "All analysis complete." emit together convey
        # completion without duplication.
        self._check(100, "Done")
        return result

    # -----------------------------------------------------------------------
    # Cross-validation
    # -----------------------------------------------------------------------

    def _cross_validate(
        self,
        x,
        y,
        presence_clean,
        feature_names,
        feature_types,
        cfg,
        cv_method,
        categorical_indices=None,
        sample_weight=None,
    ):
        """Run cross-validation and return (per-fold AUC list, first-fold split).

        Four cross-validation strategies are supported, each with its own
        canonical reference:

            cv_method == 1: Geographic K-Fold
                Spatial block CV via KMeans clustering of presence
                coordinates. Recommended for testing how well models
                trained in one region generalize to another
                (Roberts et al. 2017; Valavi et al. 2019).

            cv_method == 2: Random K-Fold (non-spatial)
                Standard k-fold partition with shuffled assignment.
                Included for direct comparability with maxent.jar /
                ENMeval / dismo workflows that use random partitioning
                by default. Tends to inflate AUC relative to spatial
                methods when presences are spatially autocorrelated
                (Roberts et al. 2017).

            cv_method == 3: Checkerboard
                Single train/test split using a checkerboard
                tessellation of the study area. Standard partition in
                ENMeval (Muscarella et al. 2014).

            cv_method == 4: Buffered Leave-One-Out
                For each presence, leave it out and remove training
                points within `buffer_dist` metres of the held-out
                point. Designed for small samples and to mitigate
                spatial autocorrelation in test scoring
                (Pearson et al. 2007; Ploton et al. 2020).

        BLOO uses the Pearson 2007 / Ploton 2020 standard of pooling
        all held-out presence predictions with a single background
        sample to compute one AUC, rather than averaging per-fold AUCs
        (which are statistically unstable when each fold has only one
        test presence sharing the entire background set).

        The first-fold split is a (train_mask, test_mask) tuple of boolean
        arrays over the full (presence + background) row index. It is
        returned so that the jackknife step can reuse the user's chosen
        spatial partitioning (Roberts et al. 2017). For BufferedLeaveOneOut
        the concept of a 'first fold' is not meaningful (each test set is a
        single point), so None is returned and jackknife falls back to a
        random 25% hold-out.

        Returns:
            (aucs, first_fold, cv_roc_fpr, cv_roc_tpr) where
                aucs: list[float] per-fold AUC values, NaNs preserved.
                first_fold: (train_mask, test_mask) or None.
                cv_roc_fpr: list of fpr arrays, one per fold (empty for
                    BLOO which produces a pooled ROC, not per-fold).
                cv_roc_tpr: list of tpr arrays, paired with cv_roc_fpr.
        """
        from sklearn.metrics import roc_auc_score
        from sklearn.metrics import roc_curve as _roc_curve

        from ..bridge import elapid_bridge as eb

        all_p = np.where(y == 1)[0]
        all_bg = np.where(y == 0)[0]

        def _build():
            return eb.make_maxent_model(
                feature_types=feature_types,
                beta_multiplier=cfg.get("beta_mult", 1.0),
                transform=cfg.get("transform", "cloglog"),
                n_hinge_features=cfg.get("n_hinge", 50),
                n_threshold_features=cfg.get("n_threshold", 50),
                use_lambdas="last",
                n_lambdas=200,
                class_weights=100,
                use_sklearn=True,
            )

        def _auc(m, xt, yt):
            try:
                return float(roc_auc_score(yt, m.predict(xt)))
            except Exception:
                # roc_auc_score raises when the test set is degenerate
                # (all-positive or all-negative) — common in small CV
                # folds. NaN propagates cleanly through nanmean so we
                # report CV AUC ± std without one bad fold killing the
                # whole metric.
                return float("nan")

        def _masks(tr_p, te_p):
            tr = np.zeros(len(y), dtype=bool)
            tr[all_p[tr_p]] = True
            tr[all_bg] = True
            te = np.zeros(len(y), dtype=bool)
            te[all_p[te_p]] = True
            te[all_bg] = True
            return tr, te

        def _to_metric_gdf(gdf):
            """Reproject the presence GeoDataFrame to EPSG:6933 if its
            CRS is geographic (degrees).

            EPSG:6933 (World Equidistant Cylindrical) is area-
            preserving and uses metres — the same CRS used in
            priority_sites.py for distance filters. Reprojecting here
            lets the user enter `Grid size (m)` and `Buffer (m)` as
            literal metres regardless of whether the presence layer
            arrived in EPSG:4326 (degrees) or some other CRS.

            We only reproject when needed:
              - geographic (lat/lon) → EPSG:6933
              - already projected (any metric CRS) → leave as is

            elapid's checkerboard_split and BufferedLeaveOneOut both
            consume coordinates in the GeoDataFrame's native CRS, so
            this single reprojection step is sufficient.
            """
            try:
                src_crs = getattr(gdf, "crs", None)
                if src_crs is None:
                    return gdf
                if bool(src_crs.is_geographic):
                    return gdf.to_crs("EPSG:6933")
                return gdf
            except Exception:
                # If the CRS lookup or reprojection fails (mis-encoded
                # CRS, missing PROJ database, etc.), fall back to the
                # original GeoDataFrame — better an approximate split
                # than a hard failure mid-run.
                return gdf

        aucs = []
        first_fold = None
        # Per-fold ROC traces. We collect (fpr, tpr) for every fold so the
        # Results-tab ROC panel can render the academic-standard plot
        # (each fold's curve in the background, mean curve highlighted)
        # instead of a single training ROC. Empty for BLOO because BLOO
        # produces one pooled AUC, not per-fold ROC traces.
        cv_roc_fpr = []
        cv_roc_tpr = []

        if cv_method == 1:  # Geographic K-Fold
            gkf = eb.make_geographic_kfold(n_splits=cfg.get("n_folds", 5))
            pts = presence_clean.reset_index(drop=True)
            for fi, (tr_p, te_p) in enumerate(gkf.split(pts)):
                if self._cancelled:
                    break
                if len(te_p) == 0:
                    continue
                tr, te = _masks(tr_p, te_p)
                if first_fold is None:
                    first_fold = (tr, te)
                sw_tr = sample_weight[tr] if sample_weight is not None else None
                m = _build()
                m.fit(
                    x[tr],
                    y[tr],
                    sample_weight=sw_tr,
                    categorical=categorical_indices or None,
                    labels=feature_names,
                )
                # Compute predictions once and derive both AUC and the
                # ROC trace from them — saves a second pass through
                # the model.
                try:
                    preds = m.predict(x[te])
                    auc = float(roc_auc_score(y[te], preds))
                    fpr_k, tpr_k, _ = _roc_curve(y[te], preds)
                    cv_roc_fpr.append(fpr_k.tolist())
                    cv_roc_tpr.append(tpr_k.tolist())
                except Exception:
                    auc = float("nan")
                aucs.append(auc)
                self.log.emit(f"  Fold {fi + 1}: {len(te_p):>3d} test presences, AUC = {auc:.4f}")

        elif cv_method == 2:  # Random K-Fold (Phillips 2006)
            # Non-spatial k-fold partition. Included for direct
            # comparability with maxent.jar / ENMeval / dismo
            # workflows that use random k-fold by default. We use
            # sklearn's KFold(shuffle=True, random_state=…) so the
            # partition is deterministic given cfg["random_seed"]
            # (else the seed) — matches typical R workflow expectations
            # of reproducible random folds. Note: Roberts et al. 2017
            # show this partition tends to inflate AUC relative to
            # spatial methods when presences are autocorrelated, hence
            # the warning in the UI tooltip.
            from sklearn.model_selection import KFold

            n_splits = int(cfg.get("n_folds", 5))
            # cfg["random_seed"] is None when the user unchecked
            # "Fix random seed". sklearn's random_state=None falls
            # back to the global numpy RNG, which uses OS entropy
            # — exactly the behavior the unchecked state promises.
            seed_cfg = cfg.get("random_seed")
            seed = int(seed_cfg) if seed_cfg is not None else None
            kf = KFold(n_splits=n_splits, shuffle=True, random_state=seed)
            pts = presence_clean.reset_index(drop=True)
            for fi, (tr_p, te_p) in enumerate(kf.split(pts)):
                if self._cancelled:
                    break
                if len(te_p) == 0:
                    continue
                tr, te = _masks(tr_p, te_p)
                if first_fold is None:
                    first_fold = (tr, te)
                sw_tr = sample_weight[tr] if sample_weight is not None else None
                m = _build()
                m.fit(
                    x[tr],
                    y[tr],
                    sample_weight=sw_tr,
                    categorical=categorical_indices or None,
                    labels=feature_names,
                )
                try:
                    preds = m.predict(x[te])
                    auc = float(roc_auc_score(y[te], preds))
                    fpr_k, tpr_k, _ = _roc_curve(y[te], preds)
                    cv_roc_fpr.append(fpr_k.tolist())
                    cv_roc_tpr.append(tpr_k.tolist())
                except Exception:
                    auc = float("nan")
                aucs.append(auc)
                self.log.emit(f"  Fold {fi + 1}: {len(te_p):>3d} test presences, AUC = {auc:.4f}")

        elif cv_method == 3:  # Checkerboard (single train/test split)
            # v0.1.7: reproject to EPSG:6933 (metres) so the user's
            # Grid size value is consumed as metres regardless of
            # whether presence_clean arrived in EPSG:4326 (degrees)
            # or a projected CRS. Default 50000 = 50 km, a reasonable
            # checkerboard scale for continent-wide SDMs.
            pts = _to_metric_gdf(presence_clean.copy())
            pts["_oi"] = np.arange(len(presence_clean))
            tr_pts, te_pts = eb.checkerboard_split(pts, grid_size=cfg.get("grid_size", 50000.0))
            if len(te_pts) > 0:
                tr, te = _masks(tr_pts["_oi"].values, te_pts["_oi"].values)
                first_fold = (tr, te)
                sw_tr = sample_weight[tr] if sample_weight is not None else None
                m = _build()
                m.fit(
                    x[tr],
                    y[tr],
                    sample_weight=sw_tr,
                    categorical=categorical_indices or None,
                    labels=feature_names,
                )
                try:
                    preds = m.predict(x[te])
                    auc = float(roc_auc_score(y[te], preds))
                    fpr_k, tpr_k, _ = _roc_curve(y[te], preds)
                    cv_roc_fpr.append(fpr_k.tolist())
                    cv_roc_tpr.append(tpr_k.tolist())
                except Exception:
                    auc = float("nan")
                aucs.append(auc)
                self.log.emit(f"  Split: {len(te_pts):>3d} test presences, AUC = {auc:.4f}")

        elif cv_method == 4:  # Buffered LOO
            # Pearson 2007 / Ploton 2020 standard: pool predictions across
            # all single-point folds and compute one AUC over the pooled
            # presences + background, instead of averaging per-fold AUCs
            # (which are statistically unstable because each fold has only
            # one test presence and shares the entire background set).
            #
            # v0.1.7: `buffer_dist` is always interpreted as metres,
            # regardless of the source CRS. We achieve this by
            # reprojecting the presence GeoDataFrame to EPSG:6933
            # (an area-preserving metric CRS) before handing it to
            # elapid's BufferedLeaveOneOut. The 50,000 default is
            # 50 km — illustrative only; an appropriate value depends
            # on the species' dispersal range and the spatial
            # autocorrelation length of the covariates (Roberts et
            # al. 2017; Ploton et al. 2020).
            bloo = eb.make_buffered_loo(distance=cfg.get("buffer_dist", 50000))
            pts = _to_metric_gdf(presence_clean.reset_index(drop=True).copy())
            pts["_y"] = 1  # for class_label

            test_pres_pred = []  # predictions for held-out presences
            n_folds_run = 0
            # Skip BLOO folds whose training set has fewer than this many
            # presences. With n_train < 5 the fitted model is unstable and
            # tends to produce extreme predictions that distort the pooled
            # AUC. The default of 5 is a conservative compromise: large
            # enough to keep folds meaningful but small enough that small-
            # sample-size species (n_presence ≈ 20) still produce some
            # valid folds. Exposed as a hidden config key for advanced
            # users; not surfaced in the UI to keep the Parameters tab
            # focused on the published Maxent options.
            min_train_pres = cfg.get("bloo_min_train_pres", 5)

            # Estimate total folds for progress reporting. BLOO splits
            # produce one fold per presence point, so we can bound the
            # progress bar to that range — useful because BLOO can
            # take minutes on a 100+ presence species (one retrain
            # per held-out point).
            n_total_folds = max(1, len(presence_clean))
            self.log.emit(f"  BLOO: {n_total_folds} folds total (one retrain per presence point)")

            fold_idx = 0
            for tr_p, te_p in bloo.split(pts, class_label="_y"):
                if self._cancelled:
                    break
                fold_idx += 1
                # Progress: BLOO occupies the 50–75% range. Steps go
                # 50 → 75 across the folds. We update on every fold
                # so a long BLOO run doesn't look frozen.
                pct = 50 + int((fold_idx / n_total_folds) * 25)
                self.progress.emit(
                    pct,
                    f"BLOO fold {fold_idx}/{n_total_folds}",
                )
                if len(tr_p) < min_train_pres:
                    continue
                tr, te = _masks(tr_p, te_p)
                try:
                    sw_tr = sample_weight[tr] if sample_weight is not None else None
                    m = _build()
                    m.fit(
                        x[tr],
                        y[tr],
                        sample_weight=sw_tr,
                        categorical=categorical_indices or None,
                        labels=feature_names,
                    )
                    # Held-out presence prediction(s); single value per fold
                    test_pres_pred.append(m.predict(x[all_p[te_p]]))
                    n_folds_run += 1
                except Exception:
                    # A single fold can fail for many reasons in
                    # spatial CV: too few presences in the training
                    # split, an all-zero categorical column after
                    # spatial blocking, glmnet convergence failure on
                    # a degenerate fold. We skip the fold and report
                    # the metric over the folds that succeeded; the
                    # final CV summary explicitly logs n_folds_run.
                    pass

            if n_folds_run > 0 and len(test_pres_pred) > 0:
                # Pool: held-out presence predictions + a single background
                # prediction sample using the model trained on all data
                # is not faithful to LOO; instead, we use background
                # predictions averaged across folds. Standard practice
                # (Pearson 2007) is to score each held-out presence against
                # the same background pool, so we approximate by predicting
                # background once with the last fitted model — equivalent up
                # to small variance to refitting on all data because all
                # folds drop only 1 point.
                bg_pred = m.predict(x[all_bg])
                pooled_pres = np.concatenate(test_pres_pred)
                pooled_y = np.concatenate([np.ones(len(pooled_pres)), np.zeros(len(bg_pred))])
                pooled_pred = np.concatenate([pooled_pres, bg_pred])
                try:
                    pooled_auc = float(roc_auc_score(pooled_y, pooled_pred))
                except Exception:
                    # Same degenerate-set guard as _auc above; this is
                    # the per-presence pooled AUC variant used by
                    # Buffered Leave-One-Out CV.
                    pooled_auc = float("nan")
                aucs.append(pooled_auc)
                self.log.emit(f"  Pooled AUC = {pooled_auc:.4f} ({n_folds_run} folds run)")
            # first_fold stays None for BLOO -> jackknife falls back to
            # random hold-out per decision #3-a-i.

        return aucs, first_fold, cv_roc_fpr, cv_roc_tpr

    # -----------------------------------------------------------------------
    # -----------------------------------------------------------------------
    # Jackknife
    # -----------------------------------------------------------------------

    def _jackknife(
        self,
        full_model,
        x,
        y,
        feature_names,
        feature_types,
        cfg,
        first_fold,
        categorical_indices=None,
        sample_weight=None,
        cv_aucs=None,
    ):
        """Jackknife variable importance with train+test AUC reporting.

        Returns:
            (results, full_train_auc, full_test_auc) where
              results: list of per-variable dicts with only/without/drop AUCs
              full_train_auc: full-model AUC on the whole dataset
              full_test_auc:  full-model AUC averaged across folds (or
                              single hold-out when CV is disabled / BLOO)

        When ``cv_aucs`` is provided (the list of per-fold held-out AUCs
        already computed in § 5 Cross-validation), this method uses it
        directly as ``full_test_auc`` instead of refitting the full
        model fold-by-fold. This (a) guarantees the reported
        full-model reference matches the headline CV AUC exactly —
        no second fold-split, no second retrain, no rounding-error
        divergence — and (b) saves ``n_folds`` full-model fits on the
        critical path. The per-variable ``only(v)`` / ``without(v)``
        retrains still need to happen, of course.

        For every variable v we fit two reduced models per fold:
            only(v):    model trained on v alone
            without(v): model trained on all variables except v
        and report mean AUC on both training and held-out (test) splits
        across folds. Reporting an across-fold mean (rather than the
        first fold only) brings the jackknife panel onto the same
        statistical footing as the ROC panel, which already reports the
        K-fold mean AUC. Phillips' maxent.jar manual likewise specifies
        that "values shown are averages over replicate runs".

        Reporting train AND test AUC (not training only) follows the
        warning of Lobo, Jimenez-Valverde & Real (2008) that training-set
        AUC is a misleading measure of SDM generalization.

        When the user has enabled cross-validation (Geographic K-Fold,
        Random K-Fold, Checkerboard), the same fold partitioning is
        reused here so the jackknife respects the user's chosen CV
        scheme. For BLOO and the no-CV case we fall back to a single
        stratified random 25% hold-out (Maxent's `randomtestpoints=25`
        default).

        References:
            Phillips, S.J., Anderson, R.P., & Schapire, R.E. (2006).
                Maximum entropy modeling of species geographic
                distributions. Ecological Modelling, 190, 231-259.
            Phillips, S.J., Anderson, R.P., Dudik, M., Schapire, R.E., &
                Blair, M.E. (2017). Opening the black box: an
                open-source release of Maxent. Ecography, 40, 887-893.
            Lobo, J.M., Jimenez-Valverde, A., & Real, R. (2008).
                AUC: a misleading measure of the performance of
                predictive distribution models. Global Ecology and
                Biogeography, 17, 145-151.
            Roberts, D.R., et al. (2017). Cross-validation strategies
                for data with temporal, spatial, hierarchical, or
                phylogenetic structure. Ecography, 40, 913-929.

        Args:
            full_model:    already-trained model on all variables.
            x, y:          design matrix and labels (presence + background).
            feature_names: list of variable names matching x's columns.
            feature_types: feature types used for the full model.
            cfg:           run configuration.
            first_fold:    (train_mask, test_mask) tuple from the user's
                CV partitioning (used only as a fallback hint when fold
                regeneration fails — the new fold loop generates its
                own splits via ``cfg["cv_method"]`` for full coverage).

        Returns:
            list of per-variable dicts with 7 keys (variable + 4 AUC + 2 drop).
            test_* keys are NaN when n_presence < 10 (decision #3-e).
        """
        from sklearn.metrics import roc_auc_score
        from sklearn.model_selection import KFold, train_test_split

        from ..bridge import elapid_bridge as eb

        def _auc(m, xt, yt):
            try:
                return float(roc_auc_score(yt, m.predict(xt)))
            except Exception:
                return float("nan")

        def _build():
            return eb.make_maxent_model(
                feature_types=feature_types,
                beta_multiplier=cfg.get("beta_mult", 1.0),
                transform=cfg.get("transform", "cloglog"),
                n_hinge_features=cfg.get("n_hinge", 50),
                n_threshold_features=cfg.get("n_threshold", 50),
                use_lambdas="last",
                n_lambdas=200,
                class_weights=100,
                use_sklearn=True,
            )

        # ── Build the list of folds we'll average across ─────────────────
        # When CV is configured we reuse the same partitioning the main
        # run uses, so jackknife metrics mirror the ROC panel's mean. For
        # BLOO and no-CV we fall back to a single 25% hold-out — averaging
        # across one fold is just that fold's value, but the code path
        # remains uniform.
        n_presence = int(y.sum())
        report_test = n_presence >= 10
        cv_method = int(cfg.get("cv_method", 0))
        n_folds_cfg = int(cfg.get("n_folds", 5))
        # None when the user unchecked "Fix random seed" — passes
        # through to sklearn / train_test_split as random_state=None
        # which means "draw from OS entropy", the behavior the
        # unchecked state promises.
        seed_cfg = cfg.get("random_seed")
        seed = int(seed_cfg) if seed_cfg is not None else None

        all_p = np.where(y == 1)[0]
        all_bg = np.where(y == 0)[0]

        fold_masks = []
        if not report_test:
            tr = np.ones(len(y), dtype=bool)
            te = np.zeros(len(y), dtype=bool)
            fold_masks.append((tr, te))
            self.log.emit(f"  Jackknife: n_presence={n_presence} < 10 — reporting train AUC only.")
        elif cv_method in (1, 2):
            # 1 = Geographic K-Fold, 2 = Random K-Fold
            try:
                if cv_method == 1:
                    presence_clean = cfg.get("_presence_clean")
                    if presence_clean is not None:
                        splitter = eb.make_geographic_kfold(n_splits=n_folds_cfg)
                        src = presence_clean.reset_index(drop=True)
                    else:
                        splitter = KFold(n_splits=n_folds_cfg, shuffle=True, random_state=seed)
                        src = np.arange(len(all_p))
                else:
                    splitter = KFold(n_splits=n_folds_cfg, shuffle=True, random_state=seed)
                    src = np.arange(len(all_p))
                for tr_p, te_p in splitter.split(src):
                    if len(te_p) == 0:
                        continue
                    tr = np.zeros(len(y), dtype=bool)
                    tr[all_p[tr_p]] = True
                    tr[all_bg] = True
                    te = np.zeros(len(y), dtype=bool)
                    te[all_p[te_p]] = True
                    te[all_bg] = True
                    fold_masks.append((tr, te))
            except Exception as e:
                # Fall back to a single hold-out if the splitter blows
                # up (e.g. Geographic K-Fold with degenerate coords).
                self.log.emit(
                    f"  Jackknife: fold setup failed "
                    f"({type(e).__name__}); falling back to 25% hold-out."
                )
                fold_masks = []
            if fold_masks:
                # (No "averaging across N folds" line here — the Block 4
                # header above already states the fold count.)
                pass

        if not fold_masks:
            # Either CV is disabled / BLOO / Checkerboard, or fold
            # construction failed above. Single 25% hold-out.
            if first_fold is not None:
                # Reuse the existing first-fold split when caller passed
                # one — keeps continuity for BLOO/Checkerboard runs.
                fold_masks.append(first_fold)
            else:
                tr_p, te_p = train_test_split(
                    np.arange(len(all_p)), test_size=0.25, random_state=seed
                )
                tr = np.zeros(len(y), dtype=bool)
                tr[all_p[tr_p]] = True
                tr[all_bg] = True
                te = np.zeros(len(y), dtype=bool)
                te[all_p[te_p]] = True
                te[all_bg] = True
                fold_masks.append((tr, te))
                self.log.emit(f"  Jackknife: random 25% hold-out ({len(te_p)} test presences).")

        len(fold_masks)

        # ── Full-model reference AUCs ────────────────────────────────────
        # Train AUC is computed once on the whole dataset (matches the
        # headline "Training AUC" emitted in § 2). Test AUC re-uses
        # the per-fold cv_aucs from § 5 when available, so:
        #   • the value reported here matches the headline CV AUC
        #     line-for-line (no second fold-split, no rounding-error
        #     divergence);
        #   • we save n_folds full-model retrains on the critical path.
        # When cv_aucs is None (CV is disabled / BLOO / Checkerboard /
        # CV fold setup failed), we fall back to retraining the full
        # model on each jackknife fold the way the v0.1.6 path did.
        full_train_auc = _auc(full_model, x, y)
        if cv_aucs is not None:
            valid_cv = [a for a in cv_aucs if not np.isnan(a)]
            full_test_auc = float(np.mean(valid_cv)) if valid_cv else float("nan")
        else:
            full_test_aucs = []
            for tr, te in fold_masks:
                if not (report_test and te.any()):
                    continue
                sw_tr = sample_weight[tr] if sample_weight is not None else None
                try:
                    m = _build()
                    m.fit(
                        x[tr],
                        y[tr],
                        sample_weight=sw_tr,
                        categorical=categorical_indices or None,
                        labels=feature_names,
                    )
                    full_test_aucs.append(_auc(m, x[te], y[te]))
                except Exception:
                    full_test_aucs.append(float("nan"))
            full_test_auc = float(np.nanmean(full_test_aucs)) if full_test_aucs else float("nan")
        self.log.emit(
            f"  Reference: full model — train AUC = {full_train_auc:.4f}"
            + (f", mean held-out test AUC = {full_test_auc:.4f}" if report_test else "")
        )

        # ── Per-variable jackknife loop ──────────────────────────────────
        # Emit a one-time legend so the per-variable lines below
        # (only_tr=, only_te=, without_tr=, without_te=) stay short
        # while keeping the meaning obvious on first read.
        # Wording is method-neutral: K-Fold runs average across N
        # folds, Checkerboard runs use the single train/test split,
        # BLOO / None use a 25% hold-out — so "evaluated on the test
        # split(s)" covers all four without misleading the reader.
        self.log.emit("  Per-variable AUCs (tr = train AUC, te = held-out test AUC):")
        results = []
        n = len(feature_names)
        cat_set = set(categorical_indices or [])

        for i, name in enumerate(feature_names):
            if self._cancelled:
                break
            pct = 75 + int(((i + 1) / n) * 20)
            self.progress.emit(pct, f"Jackknife [{i + 1}/{n}]: {name}")

            x_only = x[:, [i]]
            cols = [j for j in range(n) if j != i]
            x_wo = x[:, cols]
            names_w = [feature_names[j] for j in cols]

            cat_only = [0] if i in cat_set else None
            cat_wo = [cols.index(j) for j in categorical_indices or [] if j != i] or None

            # Per-fold metric arrays --------------------------------------
            only_tr_aucs = []
            only_te_aucs = []
            wo_tr_aucs = []
            wo_te_aucs = []
            only_skip_reason = None
            without_skip_reason = None

            for tr, te in fold_masks:
                sw_tr = sample_weight[tr] if sample_weight is not None else None

                # Only this variable -----------------------------------
                if i in cat_set:
                    # Categorical-only: dummy-column workaround for
                    # elapid 1.x (LinearTransformer needs ≥1 column).
                    n_rows = x_only.shape[0]
                    dummy = np.zeros((n_rows, 1), dtype=x_only.dtype)
                    x_aug = np.hstack([x_only, dummy])
                    cat_aug = [0]
                    labels_aug = [name, "_dummy_zero_"]
                    try:
                        m = _build()
                        m.fit(
                            x_aug[tr],
                            y[tr],
                            sample_weight=sw_tr,
                            categorical=cat_aug,
                            labels=labels_aug,
                        )
                        tr_auc = _auc(m, x_aug[tr], y[tr])
                        te_auc = (
                            _auc(m, x_aug[te], y[te]) if report_test and te.any() else float("nan")
                        )
                        if abs(tr_auc - 0.5) < 0.05:
                            if only_skip_reason is None:
                                only_skip_reason = (
                                    f"only-* skipped: dummy-column "
                                    f"workaround produced a near-random "
                                    f"model (train AUC={tr_auc:.3f}); "
                                    f"maxnet's lasso regularisation "
                                    f"collapsed the OneHot weights. "
                                    f"Lower the regularization "
                                    f"multiplier or read importance "
                                    f"from the without-* row."
                                )
                            continue
                        only_tr_aucs.append(tr_auc)
                        only_te_aucs.append(te_auc)
                    except Exception as e:
                        if only_skip_reason is None:
                            only_skip_reason = f"only-* failed: {type(e).__name__}: {str(e)[:200]}"
                else:
                    try:
                        m = _build()
                        m.fit(
                            x_only[tr],
                            y[tr],
                            sample_weight=sw_tr,
                            categorical=cat_only,
                            labels=[name],
                        )
                        only_tr_aucs.append(_auc(m, x_only[tr], y[tr]))
                        if report_test and te.any():
                            only_te_aucs.append(_auc(m, x_only[te], y[te]))
                    except Exception as e:
                        if only_skip_reason is None:
                            only_skip_reason = f"only-* failed: {type(e).__name__}: {str(e)[:200]}"

                # Without this variable --------------------------------
                try:
                    m = _build()
                    m.fit(x_wo[tr], y[tr], sample_weight=sw_tr, categorical=cat_wo, labels=names_w)
                    wo_tr_aucs.append(_auc(m, x_wo[tr], y[tr]))
                    if report_test and te.any():
                        wo_te_aucs.append(_auc(m, x_wo[te], y[te]))
                except Exception as e:
                    if without_skip_reason is None:
                        without_skip_reason = (
                            f"without-* failed: {type(e).__name__}: {str(e)[:200]}"
                        )

            def _mean(arr):
                valid = [v for v in arr if not np.isnan(v)]
                return float(np.mean(valid)) if valid else float("nan")

            only_train_auc = _mean(only_tr_aucs)
            only_test_auc = _mean(only_te_aucs)
            without_train_auc = _mean(wo_tr_aucs)
            without_test_auc = _mean(wo_te_aucs)

            train_drop = full_train_auc - without_train_auc
            test_drop = full_test_auc - without_test_auc if report_test else float("nan")

            results.append(
                {
                    "variable": name,
                    "only_train_auc": round(only_train_auc, 4),
                    "only_test_auc": round(only_test_auc, 4),
                    "without_train_auc": round(without_train_auc, 4),
                    "without_test_auc": round(without_test_auc, 4),
                    "train_drop_without": round(train_drop, 4),
                    "test_drop_without": round(test_drop, 4),
                }
            )
            # Compute the variable-name column width once per call
            # so all per-variable lines line up. We pad to the longest
            # feature name in the current run.
            name_w = max(len(nm) for nm in feature_names)
            self.log.emit(
                f"    {name:<{name_w}}  "
                f"only_tr = {only_train_auc:.4f}  "
                f"only_te = {only_test_auc:.4f}  "
                f"without_tr = {without_train_auc:.4f}  "
                f"without_te = {without_test_auc:.4f}"
            )
            if only_skip_reason:
                self.log.emit(f"      {only_skip_reason}")
            if without_skip_reason:
                self.log.emit(f"      {without_skip_reason}")

        return results, full_train_auc, full_test_auc

    def _save_xlsx(self, result: dict, cfg: dict):
        """Multi-sheet styled XLSX combining everything a researcher
        needs to cite the run, in a layout that mirrors the
        Supplementary-Table conventions of academic papers (Times New
        Roman 11pt, top/bottom rules with a thin header rule, no
        vertical borders, bold first column, row spacing of ~1.4×).

        Sheets:
            Overview          — experiment metadata in 3-column
                                Category / Item / Value layout
            Variables         — variable inventory + training range
            Cross-validation  — per-fold AUC + mean ± std
            Jackknife         — per-variable only/without × train/test AUC
            Priority Sites    — present only if the priority-sites step
                                wrote thresholds into the run

        Each row carries enough self-contained context that a paper's
        Supplementary section can reference an entire sheet without
        further reformatting.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Alignment,
                Border,
                Font,
                Side,
            )
        except Exception as e:
            # openpyxl is in REQUIRED_PACKAGES so this should never
            # trigger in production, but degrade gracefully if it does.
            self.log.emit(
                f"  ⚠ XLSX export skipped — openpyxl unavailable ({type(e).__name__}: {e})"
            )
            return

        import datetime as _dt

        # ── Style constants ──────────────────────────────────────────────
        # Academic-paper table conventions (matches Lee et al. and
        # similar Elsevier/Springer Supplementary Tables):
        #   • Times New Roman, 11pt for body, 11pt bold for headers
        #   • Top + bottom rules thick (1.5pt), header bottom thin (0.5pt)
        #   • No vertical borders
        #   • Section title row: bold, slightly larger
        FONT_BODY = Font(name="Times New Roman", size=11)
        FONT_BOLD = Font(name="Times New Roman", size=11, bold=True)
        FONT_ITALIC = Font(name="Times New Roman", size=10, italic=True)
        FONT_TITLE = Font(name="Times New Roman", size=12, bold=True)

        ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

        SIDE_THICK = Side(border_style="medium", color="000000")
        SIDE_THIN = Side(border_style="thin", color="000000")

        Border(top=SIDE_THICK)
        Border(bottom=SIDE_THIN)
        Border(bottom=SIDE_THICK)

        ROW_HEIGHT = 24  # ≈ 1.6× line spacing at 11pt Times New Roman

        # ── Helpers ──────────────────────────────────────────────────────
        def _write_section_title(ws, row, text, n_cols):
            """Write a 'Table N. Description' row above the table."""
            ws.cell(row=row, column=1, value=text).font = FONT_TITLE
            ws.cell(row=row, column=1).alignment = ALIGN_LEFT
            ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=n_cols)
            ws.row_dimensions[row].height = 24

        def _write_header(ws, row, headers):
            """Write a header row with bold center-aligned text and a
            thick top rule + thin bottom rule (no left/right borders).
            """
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = FONT_BOLD
                cell.alignment = ALIGN_CENTER
                # Top thick, bottom thin
                cell.border = Border(top=SIDE_THICK, bottom=SIDE_THIN)
            ws.row_dimensions[row].height = ROW_HEIGHT

        def _write_data(ws, row, values, alignments=None, bold_cols=(), is_last=False):
            """Write one body row. `alignments` is a list of
            'left'/'center'/'right' per column; `bold_cols` indicates
            columns that should render in bold (e.g. category cells in
            the 3-column layout). When `is_last` is True we draw the
            thick bottom rule below this row.
            """
            border_kwargs = {}
            if is_last:
                border_kwargs["bottom"] = SIDE_THICK
            border = Border(**border_kwargs) if border_kwargs else None

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = FONT_BOLD if col_idx in bold_cols else FONT_BODY
                if alignments:
                    a = alignments[col_idx - 1]
                    cell.alignment = (
                        ALIGN_LEFT
                        if a == "left"
                        else ALIGN_CENTER
                        if a == "center"
                        else ALIGN_RIGHT
                        if a == "right"
                        else ALIGN_LEFT
                    )
                else:
                    cell.alignment = ALIGN_LEFT
                if border is not None:
                    cell.border = border
            ws.row_dimensions[row].height = ROW_HEIGHT

        def _write_footnote(ws, row, text, n_cols, col_widths=None):
            """Write an italic footnote spanning n_cols.

            Row height adapts to text length when wrap_text causes the
            paragraph to span multiple visual lines: an explicit
            ROW_HEIGHT is too small for footnotes longer than ~80 chars
            and causes adjacent footnote rows to visually overlap when
            they're written one after another. We estimate the wrap
            count from the merged total width (sum of col_widths in
            Excel character units, ~7px each at 11pt) and grow the
            row height to fit. ``col_widths`` is optional; without it
            we fall back to a generous default that handles 1-3 lines.
            """
            ws.cell(row=row, column=1, value=text).font = FONT_ITALIC
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )
            ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=n_cols)
            # Estimate visual line count from total merged width.
            # Excel "column width" units ≈ characters of the default
            # font at 11pt. The footnote is 10pt italic Times New Roman
            # which is actually slightly wider per character than
            # regular 11pt Calibri (the assumed baseline) due to the
            # italic obliqueness, so we use 0.9 — a more conservative
            # chars-per-unit factor than the previous 1.05. The old
            # 1.05 over-estimated chars-per-line and produced footnote
            # rows that visually clipped on the ~190-260 char footnotes
            # (Overview, Variables, Jackknife).
            total_width = sum(col_widths) if col_widths else max(60, n_cols * 18)
            chars_per_line = max(40, int(total_width * 0.9))
            n_lines = max(1, (len(text) + chars_per_line - 1) // chars_per_line)
            # Each line at 10pt italic needs ~15pt of row height
            # (10pt glyphs + leading + descender slack); add 6pt of
            # top/bottom padding so adjacent footnotes don't touch
            # each other.
            ws.row_dimensions[row].height = max(ROW_HEIGHT, 15 * n_lines + 6)

        def _autosize(ws, widths):
            from openpyxl.utils import get_column_letter

            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            # Fit-to-width on print/PDF: keeps the academic table on
            # one horizontal page when researchers paste the rendered
            # output into a manuscript Supplementary section. height=0
            # means "as many vertical pages as needed".
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.6
            ws.page_margins.bottom = 0.6

        # ── Gather context ───────────────────────────────────────────────
        meta = result.get("meta", {}) or {}
        cv_labels = ["None", "Geographic K-Fold", "Random K-Fold", "Checkerboard", "Buffered LOO"]
        cv_label = cv_labels[cfg.get("cv_method", 0)]
        try:
            from .. import __version__ as _qmx_version
        except Exception:
            _qmx_version = "unknown"

        wb = Workbook()
        # Remove the default sheet; we'll add named ones in order.
        wb.remove(wb.active)

        # ════════════════════════════════════════════════════════════════
        # Sheet 1: Overview
        # ════════════════════════════════════════════════════════════════
        # Layout follows the paper's "Table 1" idiom: three columns —
        # Category / Item / Value — with the first column written only
        # on the first row of each category group (so the eye groups
        # related rows visually without sortable repetition).
        ws = wb.create_sheet("Overview")
        _write_section_title(
            ws,
            1,
            "Table 1. Experimental setup, training data, and run-level metrics.",
            n_cols=3,
        )

        # Build (category, [(item, value), ...]) groups.
        n_p = meta.get("n_presence", "")
        n_b = meta.get("n_background", "")
        full_train = result.get("full_auc")
        cv_aucs = result.get("cv_aucs", []) or []
        cv_valid = [
            a for a in cv_aucs if a is not None and not (isinstance(a, float) and np.isnan(a))
        ]
        cv_mean = round(float(np.mean(cv_valid)), 4) if cv_valid else ""
        cv_std = round(float(np.std(cv_valid)), 4) if cv_valid else ""
        full_test_jk = result.get("jk_full_test_auc")

        # None ⇒ user unchecked "Fix random seed"; record this
        # explicitly so the run is recoverable from the xlsx alone.
        seed_cfg = cfg.get("random_seed")
        seed_disp = seed_cfg if seed_cfg is not None else "random (not fixed)"

        groups = [
            (
                "Run",
                [
                    ("Generated", _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                    ("QMaxent version", _qmx_version),
                    ("Random seed", seed_disp),
                ],
            ),
            (
                "Training data",
                [
                    ("Presence points", n_p),
                    ("Background points", n_b),
                    (
                        "Categorical variables",
                        ", ".join(
                            meta.get("feature_names", [])[i]
                            for i in (meta.get("categorical_indices", []) or [])
                            if i < len(meta.get("feature_names", []) or [])
                        )
                        or "—",
                    ),
                ],
            ),
            (
                "Model",
                [
                    ("Output transform", meta.get("transform", "")),
                    ("Beta multiplier", meta.get("beta_multiplier", "")),
                    ("Feature types", ", ".join(meta.get("feature_types", []) or []) or "—"),
                    ("Distance weights", "yes" if meta.get("distance_weights") else "no"),
                ],
            ),
            (
                "Cross-validation",
                [
                    ("Method", cv_label),
                    ("Folds requested", cfg.get("n_folds", "")),
                    ("Folds completed", len(cv_valid) if cv_valid else "—"),
                ],
            ),
            (
                "Performance",
                [
                    ("Training AUC (full data)", f"{float(full_train):.4f}" if full_train else "—"),
                    (
                        "CV AUC (mean ± std)",
                        (f"{float(cv_mean):.4f} ± {float(cv_std):.4f}" if cv_valid else "—"),
                    ),
                    (
                        "Jackknife full-model test AUC",
                        f"{float(full_test_jk):.4f}" if full_test_jk else "—",
                    ),
                ],
            ),
        ]

        _write_header(ws, 2, ["Category", "Item", "Value"])
        cur = 3
        total_rows = sum(len(items) for _, items in groups)
        seen = 0
        for cat, items in groups:
            first_in_group = True
            for item, value in items:
                seen += 1
                is_last = seen == total_rows
                _write_data(
                    ws,
                    cur,
                    [cat if first_in_group else "", item, value],
                    alignments=["left", "left", "left"],
                    bold_cols=(1,),
                    is_last=is_last,
                )
                first_in_group = False
                cur += 1
        ov_widths = [22, 36, 60]
        _write_footnote(
            ws,
            cur,
            "Generated by QMaxent. CV AUC ± std is across CV folds; "
            "the jackknife full-model test AUC is averaged across the "
            "same folds and serves as the reference for the per-variable "
            "Jackknife sheet.",
            n_cols=3,
            col_widths=ov_widths,
        )
        _autosize(ws, ov_widths)

        # ════════════════════════════════════════════════════════════════
        # Sheet 2: Variables
        # ════════════════════════════════════════════════════════════════
        ws = wb.create_sheet("Variables")
        _write_section_title(
            ws,
            1,
            "Table 2. Predictor variables, types, and training-data range.",
            n_cols=4,
        )
        cat_set = set(meta.get("categorical_indices", []) or [])
        varmin = meta.get("varmin", {}) or {}
        varmax = meta.get("varmax", {}) or {}
        cat_levels = meta.get("categorical_levels", {}) or {}
        feature_names = meta.get("feature_names", []) or []

        _write_header(ws, 2, ["#", "Variable", "Type", "Range / Levels"])
        cur = 3
        n_vars = len(feature_names)
        for i, name in enumerate(feature_names):
            v_type = "categorical" if i in cat_set else "continuous"
            if v_type == "continuous":
                if name in varmin and name in varmax:
                    range_str = (
                        f"[{round(float(varmin[name]), 4)}, {round(float(varmax[name]), 4)}]"
                    )
                else:
                    range_str = "—"
            else:
                lvls = cat_levels.get(name, [])
                range_str = ", ".join(str(x) for x in lvls) if lvls else "(levels not recorded)"
            is_last = i == n_vars - 1
            _write_data(
                ws,
                cur,
                [i + 1, name, v_type, range_str],
                alignments=["center", "left", "center", "left"],
                is_last=is_last,
            )
            cur += 1
        var_widths = [6, 30, 14, 60]
        _write_footnote(
            ws,
            cur,
            "Range gives the [min, max] observed at training time for "
            "continuous variables; for categorical variables, the levels "
            "encountered are listed. Use this range to identify "
            "extrapolation regions when projecting the model.",
            n_cols=4,
            col_widths=var_widths,
        )
        _autosize(ws, var_widths)

        # ════════════════════════════════════════════════════════════════
        # Sheet 3: Cross-validation
        # ════════════════════════════════════════════════════════════════
        if cv_valid or cv_aucs:
            ws = wb.create_sheet("Cross-validation")
            _write_section_title(
                ws,
                1,
                f"Table 3. Cross-validation results ({cv_label}).",
                n_cols=2,
            )
            _write_header(ws, 2, ["Fold", "AUC"])
            cur = 3
            for i, auc in enumerate(cv_aucs, start=1):
                if auc is None or (isinstance(auc, float) and np.isnan(auc)):
                    val = "NaN"
                else:
                    val = f"{float(auc):.4f}"
                _write_data(
                    ws,
                    cur,
                    [f"Fold {i}", val],
                    alignments=["center", "center"],
                    is_last=False,
                )
                cur += 1
            # Mean ± std summary row.
            if cv_valid:
                cv_mean_s = f"{float(cv_mean):.4f}"
                cv_std_s = f"{float(cv_std):.4f}"
                _write_data(
                    ws,
                    cur,
                    ["Mean ± std", f"{cv_mean_s} ± {cv_std_s}"],
                    alignments=["center", "center"],
                    bold_cols=(1, 2),
                    is_last=True,
                )
                cur += 1
            cv_widths = [22, 30]
            _write_footnote(
                ws,
                cur,
                f"AUC computed on each fold's held-out presences "
                f"({cv_label}). Folds with degenerate test sets (e.g. "
                f"all-positive after spatial partitioning) are reported "
                f"as NaN and excluded from the mean.",
                n_cols=2,
                col_widths=cv_widths,
            )
            _autosize(ws, cv_widths)

        # ════════════════════════════════════════════════════════════════
        # Sheet 4: Jackknife
        # ════════════════════════════════════════════════════════════════
        jk = result.get("jackknife_results", []) or []
        if jk:
            ws = wb.create_sheet("Jackknife")
            _write_section_title(
                ws,
                1,
                "Table 4. Jackknife variable importance — AUC for "
                "models with only / without each variable.",
                n_cols=7,
            )

            # Sort by descending test-AUC drop (most important first)
            # so the table reads top-down by importance, matching the
            # academic convention.
            def _key(r):
                d = r.get("test_drop_without")
                if d is None or (isinstance(d, float) and np.isnan(d)):
                    d = r.get("train_drop_without", float("-inf"))
                return -float(d) if d is not None else 0.0

            jk_sorted = sorted(jk, key=_key)

            _write_header(
                ws,
                2,
                [
                    "Variable",
                    "Only — train AUC",
                    "Only — test AUC",
                    "Without — train AUC",
                    "Without — test AUC",
                    "Train AUC drop",
                    "Test AUC drop",
                ],
            )
            cur = 3
            n = len(jk_sorted)
            for i, r in enumerate(jk_sorted):

                def _v(key, r=r):  # default arg captures loop variable (ruff B023)
                    val = r.get(key)
                    if val is None or (isinstance(val, float) and np.isnan(val)):
                        return "NaN"
                    # 4-decimal fixed-point string keeps the column
                    # visually aligned (0.89 → "0.8900"), matching the
                    # academic table convention. We return a string so
                    # Excel doesn't strip trailing zeros on display.
                    return f"{float(val):.4f}"

                is_last = i == n - 1
                _write_data(
                    ws,
                    cur,
                    [
                        r.get("variable", ""),
                        _v("only_train_auc"),
                        _v("only_test_auc"),
                        _v("without_train_auc"),
                        _v("without_test_auc"),
                        _v("train_drop_without"),
                        _v("test_drop_without"),
                    ],
                    alignments=["left", "center", "center", "center", "center", "center", "center"],
                    is_last=is_last,
                )
                cur += 1

            # Optional skip-reason footnotes (only-* skipped diagnostics).
            jk_widths = [22, 16, 16, 18, 18, 14, 14]
            skipped = [
                r for r in jk_sorted if r.get("only_skip_reason") or r.get("without_skip_reason")
            ]
            if skipped:
                _write_footnote(
                    ws,
                    cur,
                    "Skipped diagnostics (NaN cells):",
                    n_cols=7,
                    col_widths=jk_widths,
                )
                cur += 1
                for r in skipped:
                    if r.get("only_skip_reason"):
                        _write_footnote(
                            ws,
                            cur,
                            f"  • {r['variable']}: {r['only_skip_reason']}",
                            n_cols=7,
                            col_widths=jk_widths,
                        )
                        cur += 1
                    if r.get("without_skip_reason"):
                        _write_footnote(
                            ws,
                            cur,
                            f"  • {r['variable']}: {r['without_skip_reason']}",
                            n_cols=7,
                            col_widths=jk_widths,
                        )
                        cur += 1
            _write_footnote(
                ws,
                cur,
                "Only — model fit using a single variable. "
                "Without — model fit using all other variables. "
                "Drop = full-model AUC − without-variable AUC; "
                "larger drops indicate variables whose unique "
                "contribution is harder to recover from the others. "
                "Each AUC is the mean across CV folds. Variables are "
                "sorted by descending test-AUC drop.",
                n_cols=7,
                col_widths=jk_widths,
            )
            _autosize(ws, jk_widths)

        # ════════════════════════════════════════════════════════════════
        # Sheet 5: Permutation importance (added v0.1.3)
        # ════════════════════════════════════════════════════════════════
        # Same metric reported by maxent.jar as "permutation importance"
        # in its HTML output. Normalized as percentage of total to match
        # the maxent.jar convention and enable direct ranking comparison
        # with published Maxent SDM studies (e.g. Lee et al. 2025
        # Table 4 column 4).
        pi_rows = result.get("permutation_results", []) or []
        if pi_rows:
            ws = wb.create_sheet("Permutation importance")
            n_repeats = result.get("permutation_n_repeats", "?")
            pi_source = result.get("permutation_source", "training set")
            _write_section_title(
                ws,
                1,
                f"Table 5. Permutation importance — AUC drop when each "
                f"variable's values are randomly shuffled "
                f"(n_repeats={n_repeats}, evaluated on {pi_source}).",
                n_cols=4,
            )
            _write_header(
                ws,
                2,
                [
                    "Variable",
                    "Mean importance",
                    f"Std (n={n_repeats})",
                    "Normalized (%)",
                ],
            )
            cur = 3
            n_pi = len(pi_rows)
            for i, r in enumerate(pi_rows):
                _write_data(
                    ws,
                    cur,
                    [
                        r.get("variable", ""),
                        f"{r.get('importance_mean', 0.0):.4f}",
                        f"{r.get('importance_std', 0.0):.4f}",
                        f"{r.get('importance_pct', 0.0):.2f}",
                    ],
                    alignments=["left", "center", "center", "center"],
                    is_last=(i == n_pi - 1),
                )
                cur += 1
            pi_widths = [22, 18, 18, 16]
            _write_footnote(
                ws,
                cur,
                "Permutation importance = AUC drop after shuffling each "
                "variable's values, computed via sklearn's "
                "permutation_importance (n_repeats independent shuffles). "
                "Normalized values sum to 100% and match the convention "
                "of maxent.jar's permutation importance output. Caveat "
                "(Strobl et al. 2007; Hooker & Mentch 2019): this metric "
                "can underestimate the importance of features that are "
                "correlated with other features in the model; interpret "
                "alongside the jackknife sheet rather than alone. "
                "Variables sorted by descending mean importance.",
                n_cols=4,
                col_widths=pi_widths,
            )
            _autosize(ws, pi_widths)

        # ════════════════════════════════════════════════════════════════
        # Sheet 6: Priority Sites (optional — present when threshold info
        # is in meta from a prior priority-sites run)
        # ════════════════════════════════════════════════════════════════
        ps = meta.get("priority_thresholds", None)
        if ps:
            ws = wb.create_sheet("Priority Sites")
            _write_section_title(
                ws,
                1,
                "Table 6. Priority Sites threshold values from the training data.",
                n_cols=2,
            )
            _write_header(ws, 2, ["Threshold method", "Value"])
            cur = 3
            items = list(ps.items())
            n_it = len(items)
            for i, (k, v) in enumerate(items):
                try:
                    v_disp = round(float(v), 6)
                except Exception:
                    v_disp = v
                _write_data(
                    ws,
                    cur,
                    [k, v_disp],
                    alignments=["left", "center"],
                    is_last=(i == n_it - 1),
                )
                cur += 1
            ps_widths = [28, 22]
            _write_footnote(
                ws,
                cur,
                "Thresholds derived from the model's predictions on the "
                "training presences: MTP — minimum training presence; "
                "T10 — 10th-percentile training presence; "
                "MaxSSS — value maximizing sensitivity + specificity. "
                "Used to define the lower bound of the high-suitability "
                "band in Validation-mode priority site selection.",
                n_cols=2,
                col_widths=ps_widths,
            )
            _autosize(ws, ps_widths)

        wb.save(cfg["output_xlsx"])
        self.log.emit(f"  → Results XLSX saved: {cfg['output_xlsx']}")
